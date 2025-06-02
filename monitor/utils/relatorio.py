# monitor/utils/relatorio.py
import os
import json
import logging
from datetime import datetime, timedelta, date
from collections import defaultdict, Counter
from typing import Dict, List, Optional, Tuple
import re
from urllib.parse import urlparse

from django.conf import settings
from django.db.models import Count, Q, Max, Min, Avg, F, Case, When, Value, CharField
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.drawing.image import Image

from monitor.models import Documento, NormaVigente, TermoMonitorado, RelatorioGerado

# Importe a biblioteca da Anthropic
import anthropic # Importe a biblioteca da Anthropic
import time # Para retries

logger = logging.getLogger(__name__)

# Remova referências a MISTRAL_API_KEY_RELATORIO e MISTRAL_API_URL_RELATORIO

class ClaudeRelatorioAdapter:
    def __init__(self):
        try:
            api_key = getattr(settings, 'ANTHROPIC_API_KEY', os.environ.get("ANTHROPIC_API_KEY"))
            if not api_key:
                raise ValueError("Chave da API Anthropic (Claude) não encontrada.")
            self.client = anthropic.Anthropic(api_key=api_key)
        except Exception as e:
            logger.error(f"Erro ao inicializar o cliente Anthropic Claude no RelatorioAdapter: {e}")
            self.client = None
        # Para relatórios consolidados, um modelo mais capaz é melhor.
        # claude-3-sonnet-20240229 ou claude-3-5-sonnet-20240620 são boas opções
        self.default_model = "claude-3-5-sonnet-20240620"
        self.default_temperature = 0.3
        self.default_max_tokens = 3000 # Aumentado para o prompt detalhado

    def _call_claude(self, system_prompt: str, user_prompt: str, model: Optional[str] = None, temperature: Optional[float] = None, max_tokens: Optional[int] = None) -> Optional[str]:
        if not self.client:
            logger.error("Cliente Anthropic Claude (RelatorioAdapter) não inicializado.")
            return "Erro: Cliente Anthropic Claude (Relatório) não configurado."

        # (Mesma lógica de _call_claude da classe ClaudeProcessor, incluindo retries)
        max_retries = 3
        base_wait_time = 5  # segundos

        for attempt in range(max_retries):
            try:
                response = self.client.messages.create(
                    model=model or self.default_model,
                    system=system_prompt,
                    messages=[{"role": "user", "content": user_prompt}],
                    temperature=temperature if temperature is not None else self.default_temperature,
                    max_tokens=max_tokens or self.default_max_tokens
                )
                if response.content and isinstance(response.content, list) and len(response.content) > 0:
                    block = response.content[0]
                    if hasattr(block, 'text'): return block.text.strip()
                    logger.warning(f"Bloco de conteúdo da API Claude (Relatório) não continha 'text': {block}")
                    return "Resposta da IA (Relatório) não continha texto no bloco esperado."
                logger.warning(f"Resposta da API Claude (Relatório) não continha 'content' ou estava vazia: {response}")
                return "Resposta da IA (Relatório) com formato inesperado ou vazia."
            except anthropic.APIStatusError as e:
                logger.error(f"Erro de Status da API Anthropic (Relatório, tentativa {attempt + 1}/{max_retries}): {e.status_code} - {e.message}", exc_info=True)
                if e.status_code == 429 and attempt < max_retries - 1:
                    wait_time = base_wait_time * (2 ** attempt)
                    logger.warning(f"Rate limit (Relatório). Aguardando {wait_time}s...")
                    time.sleep(wait_time)
                    continue
                return f"Erro API Claude (Relatório): {e.status_code} - {e.message}"
            except anthropic.APIConnectionError as e:
                logger.error(f"Erro de Conexão com API Anthropic (Relatório, tentativa {attempt + 1}/{max_retries}): {e}", exc_info=True)
                if attempt < max_retries - 1: time.sleep(base_wait_time); continue
                return "Erro API Claude (Relatório): Falha na conexão."
            except Exception as e:
                logger.error(f"Erro inesperado API Claude (Relatório, tentativa {attempt + 1}/{max_retries}): {e}", exc_info=True)
                if attempt < max_retries - 1: time.sleep(base_wait_time); continue
                return f"Erro inesperado API Claude (Relatório): {str(e)}"
        return "Erro: Limite de tentativas excedido (Relatório) com API Claude."


    def gerar_insight_conjunto_documentos(self, documentos_para_analise: List[Dict]) -> str:
        if not self.client:
            return "Erro: Cliente Anthropic Claude (Relatório) não configurado para gerar insights."
        if not documentos_para_analise:
            return "Nenhum documento fornecido para análise consolidada."

        textos_concatenados = "\n\n---\n\n".join(
            f"Documento Título: {d['titulo']}\nData: {d['data_publicacao']}\nResumo IA: {d.get('resumo_ia', 'N/A')}\nPontos Críticos IA: {'; '.join(d.get('pontos_criticos_ia', []))}"
            for d in documentos_para_analise
        )
        # Limitar o tamanho do input para Claude. Claude 3.5 Sonnet tem 200K tokens.
        # 60000 caracteres são aproximadamente 15k tokens, o que é seguro.
        textos_concatenados_trunc = textos_concatenados[:60000]


        system_prompt = (
            "Você é um Consultor Tributário Estratégico e Analista Regulatório Sênior, com vasta experiência na legislação fiscal e contábil do estado do Piauí. "
            "Sua especialidade é analisar um conjunto de documentos oficiais (Diários Oficiais, Leis, Decretos, Portarias) e seus resumos, "
            "identificando implicações, padrões, tendências e fornecendo aconselhamento prático e acionável para contadores e gestores fiscais."
        )

        user_prompt_template = """Com base na seguinte coletânea de documentos fiscais/contábeis relevantes do Piauí (títulos, datas, resumos IA e pontos críticos IA), elabore uma ANÁLISE CONSOLIDADA E DETALHADA. Sua análise deve ser estruturada nas seguintes seções:

DADOS DOS DOCUMENTOS PARA ANÁLISE:
\"\"\"
{document_data}
\"\"\"

--- ANÁLISE CONSOLIDADA DETALHADA ---

**1. PRINCIPAIS TEMAS E ALTERAÇÕES LEGISLATIVAS IDENTIFICADAS:**
   - Liste de 3 a 5 temas ou categorias de alterações mais proeminentes observados no conjunto de documentos (ex: Mudanças no ICMS-ST, Novas Obrigações Acessórias para o Setor X, Alterações em Processos de Licitação, etc.).
   - Para cada tema, descreva sucintamente a natureza da mudança e, se possível, cite 1 ou 2 documentos como exemplo (apenas o título e data são suficientes para a citação).

**2. TENDÊNCIAS REGULATÓRIAS EMERGENTES PARA O PIAUÍ:**
   - Com base nos documentos, identifique de 2 a 3 tendências regulatórias que parecem estar se consolidando no estado do Piauí (ex: Maior digitalização de processos, Foco em determinado setor econômico, Aumento da fiscalização em X, Simplificação de Y).
   - Justifique brevemente cada tendência com base nas informações dos documentos.

**3. POTENCIAIS RISCOS E PONTOS DE ATENÇÃO CRÍTICOS:**
   - Destaque quaisquer riscos significativos, ambiguidades, ou aumentos de complexidade que os documentos possam introduzir para as empresas e contadores.
   - Há algum documento ou alteração que exige CAUTELA redobrada na interpretação ou implementação?

**4. RECOMENDAÇÕES ESTRATÉGICAS E AÇÕES IMEDIATAS PARA CONTADORES:**
   - Com base na sua análise, forneça de 3 a 5 recomendações práticas e acionáveis para a equipe contábil/fiscal.
   - Ex: 'Revisar urgentemente os procedimentos internos para a nova obrigação X (vide Documento Y)', 'Alertar clientes do setor Z sobre a mudança na alíquota W', 'Capacitar a equipe sobre o novo sistema ABC'.

**5. SÍNTESE DO SENTIMENTO GERAL (IMPACTO PREDOMINANTE):**
   - Considerando o conjunto de documentos, qual o sentimento/impacto predominante para o ambiente de negócios e para a prática contábil no Piauí (POSITIVO, NEGATIVO, NEUTRO, MISTO COM CAUTELA)? Justifique brevemente.

INSTRUÇÕES ADICIONAIS:
- Seja objetivo, claro e utilize linguagem técnica apropriada para contadores.
- Foque exclusivamente nas informações contidas ou inferíveis a partir dos dados dos documentos fornecidos.
- Não inclua informações externas ou especulações não fundamentadas nos textos.
- A resposta deve ser bem organizada e fácil de ler."""

        user_prompt = user_prompt_template.format(document_data=textos_concatenados_trunc)

        # max_tokens aumentado para permitir resposta detalhada
        return self._call_claude(system_prompt, user_prompt, max_tokens=3000, temperature=0.3) or "Não foi possível gerar o insight consolidado detalhado com Claude."


class AnaliseIA: # Mantém o nome, mas adapta o cliente interno
    def __init__(self, claude_client_instance: Optional[ClaudeRelatorioAdapter] = None): # Alterado
        # Alterado para usar ClaudeRelatorioAdapter
        self.claude_adapter = claude_client_instance if claude_client_instance else ClaudeRelatorioAdapter()

    # ... (métodos _identificar_fontes_principais e _contar_tipos_normas permanecem os mesmos)
    @staticmethod #relatorio.py
    def _identificar_fontes_principais(documentos: List[Documento]) -> Dict[str, int]: #relatorio.py
        logger.debug(f"Iniciando _identificar_fontes_principais para {len(documentos)} documentos.") #relatorio.py
        fontes = defaultdict(int) #relatorio.py
        try: #relatorio.py
            for doc in documentos: #relatorio.py
                fonte_val = None #relatorio.py
                if hasattr(doc, 'fonte_documento') and doc.fonte_documento: #relatorio.py
                    fonte_val = doc.fonte_documento.strip() #relatorio.py
                elif doc.url_original: #relatorio.py
                    try: #relatorio.py
                        parsed_url = urlparse(doc.url_original) #relatorio.py
                        fonte_val = parsed_url.netloc if parsed_url.netloc else doc.url_original #relatorio.py
                    except: fonte_val = doc.url_original[:70] # Fallback #relatorio.py
                
                if fonte_val: #relatorio.py
                    fontes[fonte_val] += 1 #relatorio.py
                else: #relatorio.py
                    fontes["Fonte Não Especificada"] += 1 #relatorio.py
            return dict(sorted(fontes.items(), key=lambda item: item[1], reverse=True)[:5]) #relatorio.py
        except Exception as e: #relatorio.py
            logger.error(f"Erro em _identificar_fontes_principais: {e}", exc_info=True) #relatorio.py
            return {"Erro na análise de fontes": 1} #relatorio.py

    @staticmethod #relatorio.py
    def _contar_tipos_normas(normas: List[NormaVigente]) -> Dict[str, int]: #relatorio.py
        tipos = defaultdict(int) #relatorio.py
        for norma in normas: #relatorio.py
            tipos[norma.get_tipo_display()] += 1 #relatorio.py
        return dict(sorted(tipos.items(), key=lambda item: item[1], reverse=True)) #relatorio.py


    def _gerar_insights_automaticos(self, documentos: List[Documento], normas: List[NormaVigente]) -> List[Dict[str, str]]:
        insights = [] #relatorio.py
        try: #relatorio.py
            # Seus insights baseados em contagem (mantidos) #relatorio.py
            if len(documentos) > 50: # Limite de exemplo #relatorio.py
                insights.append({ #relatorio.py
                    'titulo': 'Alto Volume de Documentos', #relatorio.py
                    'descricao': f'Sistema analisou {len(documentos)} documentos no período, indicando atividade regulatória significativa.', #relatorio.py
                    'relevancia': 'alta' #relatorio.py
                }) #relatorio.py
            
            normas_vigentes_count = len([n for n in normas if n.situacao == 'VIGENTE']) #relatorio.py
            if normas and len(normas) > 0 and normas_vigentes_count / len(normas) > 0.8: # Adicionado len(normas) > 0 #relatorio.py
                insights.append({ #relatorio.py
                    'titulo': 'Boa Taxa de Conformidade de Normas', #relatorio.py
                    'descricao': f'{(normas_vigentes_count/len(normas)*100):.1f}% das normas identificadas estão atualmente vigentes.', #relatorio.py
                    'relevancia': 'média' #relatorio.py
                }) #relatorio.py
            
            # Insight Consolidado da Claude
            docs_para_insight_ia = [] #relatorio.py
            for doc in documentos: #relatorio.py
                if doc.relevante_contabil and (hasattr(doc, 'resumo_ia') and doc.resumo_ia): #relatorio.py
                    docs_para_insight_ia.append({ #relatorio.py
                        "titulo": doc.titulo, #relatorio.py
                        "data_publicacao": doc.data_publicacao.strftime("%d/%m/%Y") if doc.data_publicacao else "N/D", #relatorio.py
                        "resumo_ia": doc.resumo_ia[:300] + "..." if doc.resumo_ia and len(doc.resumo_ia) > 300 else doc.resumo_ia, # Envia resumos curtos #relatorio.py
                        "pontos_criticos_ia": doc.metadata.get('ia_pontos_criticos', []) if hasattr(doc, 'metadata') and doc.metadata else [] #relatorio.py
                    }) #relatorio.py
            
            if docs_para_insight_ia: # Se houver documentos relevantes com resumos IA #relatorio.py
                # Alterado para usar self.claude_adapter
                insight_consolidado_ia = self.claude_adapter.gerar_insight_conjunto_documentos(docs_para_insight_ia[:10]) # Limita a 10 docs para o prompt de exemplo
                if insight_consolidado_ia and "Erro API Claude" not in insight_consolidado_ia and "Erro: Cliente Anthropic Claude" not in insight_consolidado_ia: #relatorio.py
                    insights.append({ #relatorio.py
                        'titulo': 'Análise Consolidada por IA (Claude)', # Alterado
                        'descricao': insight_consolidado_ia, #relatorio.py
                        'relevancia': 'alta' #relatorio.py
                    }) #relatorio.py
                elif insight_consolidado_ia: # Captura e mostra o erro da API, se houver
                    insights.append({
                        'titulo': 'Falha na Análise Consolidada por IA (Claude)',
                        'descricao': f"Não foi possível gerar a análise. Detalhe do erro: {insight_consolidado_ia}",
                        'relevancia': 'crítica'
                    })
            return insights #relatorio.py
        except Exception as e: #relatorio.py
            logger.error(f"Erro ao gerar insights automáticos: {e}", exc_info=True) #relatorio.py
            return [{'titulo': 'Erro nos Insights', 'descricao': str(e), 'relevancia': 'crítica'}] #relatorio.py


class RelatorioAvancado:
    CORES = { #relatorio.py
        'cabecalho': '4F81BD', #relatorio.py
        'subcabecalho': 'B8CCE4', #relatorio.py
        'destaque': 'FCD5B4', #relatorio.py
        'sucesso': 'C6EFCE', #relatorio.py
        'aviso': 'FFEB9C', #relatorio.py
        'erro': 'FFC7CE', #relatorio.py
        'neutro': 'D9D9D9', #relatorio.py
        'zebra': 'F2F2F2', #relatorio.py
    } #relatorio.py

    def __init__(self):
        self.wb = Workbook() #relatorio.py
        self.wb.remove(self.wb.active) # Remove a planilha padrão #relatorio.py
        self.estilos = self._definir_estilos() #relatorio.py
        # Alterado para usar ClaudeRelatorioAdapter
        self.claude_adapter = ClaudeRelatorioAdapter()
        # Alterado para passar a instância correta
        self.analise_ia = AnaliseIA(claude_client_instance=self.claude_adapter)

    def _definir_estilos(self) -> Dict[str, NamedStyle]:
        estilos = {}
        border_thin_side = Side(border_style="thin", color="000000")
        default_border = Border(top=border_thin_side, left=border_thin_side, right=border_thin_side, bottom=border_thin_side)

        # Estilo Cabeçalho Principal
        estilos['cabecalho_principal'] = NamedStyle(name='cabecalho_principal_relatorio') # Nome único para o estilo
        estilos['cabecalho_principal'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        estilos['cabecalho_principal'].fill = PatternFill(start_color=self.CORES['cabecalho'], end_color=self.CORES['cabecalho'], fill_type="solid")
        estilos['cabecalho_principal'].alignment = Alignment(horizontal="center", vertical="center")
        estilos['cabecalho_principal'].border = default_border


        # Estilo Título de Seção
        estilos['titulo_secao'] = NamedStyle(name='titulo_secao_relatorio')
        estilos['titulo_secao'].font = Font(name='Calibri', size=12, bold=True, color=self.CORES['cabecalho'])
        # estilos['titulo_secao'].border = default_border # Opcional para títulos de seção

        # Estilo Cabeçalho de Tabela
        estilos['cabecalho_tabela'] = NamedStyle(name='cabecalho_tabela_relatorio')
        estilos['cabecalho_tabela'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        estilos['cabecalho_tabela'].fill = PatternFill(start_color=self.CORES['subcabecalho'], end_color=self.CORES['subcabecalho'], fill_type="solid")
        estilos['cabecalho_tabela'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        estilos['cabecalho_tabela'].border = default_border

        # Estilos para células de dados
        estilos['dados_texto'] = NamedStyle(name='dados_texto_relatorio')
        estilos['dados_texto'].font = Font(name='Calibri', size=10)
        estilos['dados_texto'].alignment = Alignment(vertical="top", wrap_text=True, horizontal="left") # Adicionado horizontal left
        estilos['dados_texto'].border = default_border

        # Estilo para dados numéricos (baseado em dados_texto, mas com alinhamento diferente)
        estilos['dados_numero'] = NamedStyle(name='dados_numero_relatorio')
        estilos['dados_numero'].font = Font(name='Calibri', size=10) # Herda visualmente de dados_texto
        estilos['dados_numero'].alignment = Alignment(horizontal="right", vertical="top", wrap_text=False) # Modificação
        estilos['dados_numero'].border = default_border # Herda visualmente de dados_texto

        # Estilo para dados de data (baseado em dados_texto, mas com formato e alinhamento diferentes)
        estilos['dados_data'] = NamedStyle(name='dados_data_relatorio')
        estilos['dados_data'].font = Font(name='Calibri', size=10) # Herda visualmente de dados_texto
        estilos['dados_data'].number_format = 'DD/MM/YYYY' # Modificação
        estilos['dados_data'].alignment = Alignment(horizontal="center", vertical="top", wrap_text=False) # Modificação
        estilos['dados_data'].border = default_border # Herda visualmente de dados_texto



        for status_cor, cor_hex in self.CORES.items():

            estilos[f'fill_{status_cor}'] = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        return estilos
        
    # Dentro da classe RelatorioAvancado em relatorio.py
    def _aplicar_estilo_celula(self, celula, estilo_base_nome: str, estilo_fill_cor_nome: Optional[str] = None): # Renomeei para clareza
        """
        Aplica um estilo base (NamedStyle) e um estilo de preenchimento (PatternFill) opcional.
        estilo_fill_cor_nome deve ser a chave da cor em self.CORES, ex: 'sucesso', 'zebra'.
        """
        if estilo_base_nome in self.estilos:
            celula.style = self.estilos[estilo_base_nome] # Aplica o NamedStyle base

        # Aplica o PatternFill diretamente
        if estilo_fill_cor_nome and f'fill_{estilo_fill_cor_nome}' in self.estilos:
            # self.estilos[f'fill_{estilo_fill_cor_nome}'] já é um objeto PatternFill
            celula.fill = self.estilos[f'fill_{estilo_fill_cor_nome}'] # <--- CORREÇÃO AQUI


    def _criar_planilha_documentos_avancada(self, documentos: List[Documento]):
        """Cria a aba de Documentos com informações detalhadas e análises da IA."""
        ws = self.wb.create_sheet(title="📄 Documentos Detalhados")

        cabecalhos = [
            "ID", "Data Public.", "Título do Documento", "Tipo Documento", "Fonte Documento", "URL Original",
            "Relevante (IA)", "Justificativa Relevância (IA)", "Pontos Críticos (IA)",
            "Resumo IA", "Sentimento (IA)", "Impacto Fiscal (IA)",
            "Normas Extraídas (Regex)", "Qtd Normas",
            "Processado?", "Data Processamento", "Data Coleta"
        ]

        for col, cabecalho_texto in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=cabecalho_texto)
            # Supondo que self._aplicar_estilo_celula e self.estilos estão definidos corretamente
            self._aplicar_estilo_celula(cell, 'cabecalho_tabela')

        row_num = 2
        for doc in documentos:
            # Dados da IA (já devem estar no objeto Documento, preenchidos pelo PDFProcessor)
            justificativa_ia = doc.metadata.get('ia_relevancia_justificativa', "N/A") if hasattr(doc, 'metadata') and doc.metadata else "N/A"
            pontos_criticos_ia_lista = doc.metadata.get('ia_pontos_criticos', []) if hasattr(doc, 'metadata') and doc.metadata else []
            pontos_criticos_ia_str = "\n".join([f"- {p}" for p in pontos_criticos_ia_lista]) if pontos_criticos_ia_lista else "N/A"
            
            resumo_ia_val = getattr(doc, 'resumo_ia', doc.resumo or "N/A") # Usa resumo_ia se existir, senão resumo principal
            sentimento_ia_val = getattr(doc, 'sentimento_ia', "N/A")
            # Supondo que impacto_fiscal é um campo TextField ou CharField no modelo Documento
            # Se for um JSONField ou algo mais complexo, ajuste a forma de obter o valor.
            impacto_fiscal_ia_val = getattr(doc, 'impacto_fiscal_ia', getattr(doc, 'impacto_fiscal', "N/A"))


            normas_extraidas_obj = doc.normas_relacionadas.all()
            normas_extraidas_str = "\n".join([f"- {n.get_tipo_display()} {n.numero}/{n.ano if n.ano else ''}" for n in normas_extraidas_obj]) if normas_extraidas_obj.exists() else "Nenhuma"
            qtd_normas = normas_extraidas_obj.count()

            dados_linha = [
                doc.id,
                doc.data_publicacao, # Será formatado pelo estilo 'dados_data'
                doc.titulo,
                doc.get_tipo_documento_display() if hasattr(doc, 'tipo_documento') and doc.tipo_documento else "N/A",
                doc.fonte_documento if hasattr(doc, 'fonte_documento') and doc.fonte_documento else "N/A",
                doc.url_original,
                "SIM" if doc.relevante_contabil else "NÃO",
                justificativa_ia,
                pontos_criticos_ia_str,
                resumo_ia_val,
                sentimento_ia_val,
                impacto_fiscal_ia_val,
                normas_extraidas_str,
                qtd_normas,
                "SIM" if doc.processado else "NÃO",
                getattr(doc, 'data_processamento', None), # Será formatado se for data
                doc.data_coleta # Será formatado se for data
            ]

            estilo_fill_linha = 'fill_zebra' if row_num % 2 == 0 else None

            for col_idx, cell_value in enumerate(dados_linha, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=cell_value)
                # Aplica estilo base e de preenchimento
                # CORREÇÃO AQUI:
                if isinstance(cell_value, (datetime, date)): # Verifica se é datetime.datetime ou datetime.date
                    self._aplicar_estilo_celula(cell, 'dados_data', estilo_fill_linha)
                elif isinstance(cell_value, (int, float)) and col_idx == cabecalhos.index("Qtd Normas") + 1 : # Exemplo para Qtd Normas
                    self._aplicar_estilo_celula(cell, 'dados_numero', estilo_fill_linha)
                else: # Texto
                    self._aplicar_estilo_celula(cell, 'dados_texto', estilo_fill_linha)
                
                # Formatação condicional para relevância
                if cabecalhos[col_idx-1] == "Relevante (IA)":
                    if doc.relevante_contabil:
                        # Supondo que self.estilos['fill_sucesso'] é um PatternFill
                        cell.fill = self.estilos['fill_sucesso'] 
                    else:
                        cell.fill = self.estilos['fill_aviso'] 

            row_num += 1

        # Ajusta larguras
        larguras = [6, 12, 45, 18, 25, 35, 12, 40, 40, 50, 15, 40, 30, 10, 10, 18, 18]
        for i, largura_val in enumerate(larguras):
            if i < len(cabecalhos): # Garante que não exceda o número de colunas
                ws.column_dimensions[get_column_letter(i + 1)].width = largura_val



    def _criar_resumo_executivo(self, documentos: List[Documento], normas: List[NormaVigente]):
        """Cria a aba de Resumo Executivo com os principais insights e gráficos."""
        ws = self.wb.create_sheet(title="📊 Resumo Executivo")
        self._aplicar_estilo_celula(ws.cell(row=1, column=1, value="📊 Resumo Executivo e Insights Chave"), 'cabecalho_principal')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5) # Mescla para o título

        row_idx = 3
        ws.cell(row=row_idx, column=1, value="Principais Métricas do Período").style = self.estilos['titulo_secao']
        row_idx += 1
        
        metricas = [
            ("Total de Documentos Analisados:", len(documentos)),
            ("Documentos Relevantes (IA):", len([d for d in documentos if d.relevante_contabil])),
            ("Total de Normas Identificadas:", len(normas)),
            ("Normas Vigentes:", len([n for n in normas if n.situacao == 'VIGENTE'])),
            ("Normas Revogadas:", len([n for n in normas if n.situacao == 'REVOGADA'])),
            ("Normas A Verificar:", len([n for n in normas if n.situacao == 'A_VERIFICAR'])),
        ]
        for desc, valor in metricas:
            ws.cell(row=row_idx, column=1, value=desc).style = self.estilos['dados_texto']
            ws.cell(row=row_idx, column=2, value=valor).style = self.estilos['dados_numero']
            row_idx += 1
        row_idx +=1 # Espaço

        ws.cell(row=row_idx, column=1, value="Insights Gerados pela Análise IA").style = self.estilos['titulo_secao']
        row_idx += 1
        
        # Usando a instância de AnaliseIA que agora tem acesso ao MistralAdapter
        insights_ia = self.analise_ia._gerar_insights_automaticos(documentos, normas)
        if insights_ia:
            for insight in insights_ia:
                ws.cell(row=row_idx, column=1, value=f"Insight: {insight.get('titulo', 'N/A')}").font = Font(bold=True)
                ws.cell(row=row_idx, column=1).style = self.estilos['dados_texto']
                row_idx +=1
                ws.cell(row=row_idx, column=1, value=insight.get('descricao', 'N/A')).style = self.estilos['dados_texto']
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
                ws.row_dimensions[row_idx].height = 30 # Aumenta altura da linha para descrição
                row_idx +=1
                ws.cell(row=row_idx, column=1, value=f"Relevância Percebida: {insight.get('relevancia', 'N/A').upper()}").font = Font(italic=True)
                ws.cell(row=row_idx, column=1).style = self.estilos['dados_texto']
                row_idx += 2 # Espaço
        else:
            ws.cell(row=row_idx, column=1, value="Nenhum insight automático gerado.").style = self.estilos['dados_texto']
            row_idx +=1

        # Adicionar mais seções conforme os métodos da AnaliseIA, por exemplo:
        # - Fontes mais comuns
        # - Tipos de normas mais frequentes

        # Ajuste de colunas para esta aba
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15


    def _criar_dashboard_visual(self, documentos: List[Documento], normas: List[NormaVigente]):
        ws = self.wb.create_sheet(title="📊 Dashboard Visual")
        self._aplicar_estilo_celula(ws.cell(row=1, column=1, value="📊 Dashboard Visual de Compliance"), self.estilos['cabecalho_principal'])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        
        row_idx = 3
        ws.cell(row=row_idx, column=1, value="Distribuição de Documentos por Relevância (IA)").style = self.estilos['titulo_secao']
        row_idx += 1

        relevantes_count = len([d for d in documentos if d.relevante_contabil])
        nao_relevantes_count = len(documentos) - relevantes_count
        ws.cell(row=row_idx, column=1, value="Relevantes").style = self.estilos['dados_texto']
        ws.cell(row=row_idx, column=2, value=relevantes_count).style = self.estilos['dados_numero']
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="Não Relevantes").style = self.estilos['dados_texto']
        ws.cell(row=row_idx, column=2, value=nao_relevantes_count).style = self.estilos['dados_numero']
        row_idx += 1

        if relevantes_count + nao_relevantes_count > 0 :
            c1 = PieChart()
            labels = Reference(ws, min_col=1, min_row=row_idx -2, max_row=row_idx -1)
            data = Reference(ws, min_col=2, min_row=row_idx-2, max_row=row_idx-1)
            c1.add_data(data, titles_from_data=False)
            c1.set_categories(labels)
            c1.title = "Relevância de Documentos"
            ws.add_chart(c1, "D4") # Posição do gráfico

        row_idx += 8 # Espaço para o gráfico
        ws.cell(row=row_idx, column=1, value="Situação das Normas Identificadas").style = self.estilos['titulo_secao']
        row_idx += 1
        
        situacao_counts = Counter(n.get_situacao_display() for n in normas)
        for situacao, count in situacao_counts.items():
            ws.cell(row=row_idx, column=1, value=situacao).style = self.estilos['dados_texto']
            ws.cell(row=row_idx, column=2, value=count).style = self.estilos['dados_numero']
            row_idx+=1

        if situacao_counts:
            c2 = BarChart()
            c2.type = "col"
            c2.style = 10
            c2.title = "Situação das Normas"
            c2.y_axis.title = 'Quantidade'
            c2.x_axis.title = 'Status'
            data = Reference(ws, min_col=2, min_row=row_idx - len(situacao_counts), max_row=row_idx-1)
            cats = Reference(ws, min_col=1, min_row=row_idx - len(situacao_counts), max_row=row_idx-1)
            c2.add_data(data, titles_from_data=False)
            c2.set_categories(cats)
            ws.add_chart(c2, "D" + str(row_idx - len(situacao_counts) +2))

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


    def _salvar_relatorio(self, nome_base: str = "relatorio_compliance") -> Optional[str]:
        """Salva o workbook e retorna o caminho relativo para o modelo."""
        try:
            relatorios_dir = os.path.join(settings.MEDIA_ROOT, 'relatorios') # MEDIA_ROOT do Django settings
            os.makedirs(relatorios_dir, exist_ok=True) # Cria o diretório se não existir

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nome_arquivo_excel = f"{nome_base}_{timestamp}.xlsx"
            caminho_completo = os.path.join(relatorios_dir, nome_arquivo_excel)

            self.wb.save(caminho_completo)
            logger.info(f"Relatório Excel gerado e salvo em: {caminho_completo}")

            # Retorna o caminho relativo a MEDIA_ROOT para ser salvo no FileField
            return os.path.join('relatorios', nome_arquivo_excel)
        except Exception as e:
            logger.error(f"Erro ao salvar o relatório Excel: {e}", exc_info=True)
            return None


    def gerar_relatorio_completo(self) -> Optional[str]:
        """
        Ponto de entrada principal para gerar todas as abas do relatório.
        Recupera dados do banco e chama os métodos de criação de planilhas.
        """
        logger.info("Iniciando geração do relatório avançado de compliance...")
        try:
            # Coleta os dados (exemplo: últimos 30 dias, ou todos os relevantes não reportados)
            # Você pode adicionar filtros aqui conforme a necessidade
            trinta_dias_atras = timezone.now() - timedelta(days=30)
            documentos = Documento.objects.filter(data_coleta__gte=trinta_dias_atras).prefetch_related('normas_relacionadas').order_by('-data_publicacao')
            # Para normas, pegar todas as que foram relacionadas aos documentos do período ou todas as ativas
            normas_ids = set()
            for doc in documentos:
                for norma in doc.normas_relacionadas.all():
                    normas_ids.add(norma.id)
            normas = NormaVigente.objects.filter(id__in=list(normas_ids)).order_by('tipo', '-ano', 'numero')

            if not documentos.exists():
                logger.warning("Nenhum documento encontrado para o período. O relatório pode ficar vazio.")
                # Mesmo assim, podemos gerar um relatório vazio com cabeçalhos
            
            # Gerar as abas
            self._criar_resumo_executivo(documentos, normas)
            self._criar_planilha_documentos_avancada(documentos) # Planilha de Documentos aprimorada
            #self._criar_planilha_normas_avancada(normas)         # Planilha de Normas aprimorada
            # self._criar_analise_compliance(normas) # Revise este método
            # self._criar_tendencias_temporais(documentos) # Revise este método
            # self._criar_analise_ia(documentos, normas) # Revise ou remova se a de Documentos for suficiente
            self._criar_dashboard_visual(documentos, normas)


            # Salvar o relatório
            caminho_relativo_arquivo = self._salvar_relatorio()

            if caminho_relativo_arquivo:
                # Registrar no modelo RelatorioGerado
                RelatorioGerado.objects.create(
                    tipo='CONTABIL', # Ou um tipo mais específico como 'COMPLIANCE_AVANCADO'
                    formato='XLSX',
                    caminho_arquivo=caminho_relativo_arquivo,
                    # parametros = { ... } # Se houver parâmetros
                    # gerado_por = ... # Se tiver o usuário
                )
                logger.info("Relatório avançado gerado e registrado com sucesso.")
                return caminho_relativo_arquivo # Retorna o caminho para a view
            else:
                logger.error("Falha ao salvar o relatório gerado.")
                return None

        except Exception as e:
            logger.error(f"Erro GERAL na geração do relatório completo: {e}", exc_info=True)
            return None


# Função de fachada para ser chamada de views.py ou tasks.py
# Mantendo a compatibilidade com o nome que você usava antes.
def gerar_relatorio_contabil_avancado() -> Optional[str]:
    """Função principal de fachada para gerar o relatório avançado de compliance."""
    try:
        gerador = RelatorioAvancado()
        caminho_arquivo = gerador.gerar_relatorio_completo()
        return caminho_arquivo
    except Exception as e:
        logger.error(f"Erro ao chamar RelatorioAvancado().gerar_relatorio_completo(): {e}", exc_info=True)
        return None