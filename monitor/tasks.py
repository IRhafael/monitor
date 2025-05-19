import os
import logging
from datetime import datetime, timedelta
from celery import shared_task
from django.conf import settings
from django.core.mail import send_mail
from django.utils import timezone
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from monitor.utils.diario_scraper import DiarioOficialScraper  # ✅ correto
from monitor.utils.sefaz_integracao import IntegradorSEFAZ
from monitor.utils.sefaz_scraper import  SEFAZScraper
from monitor.utils import PDFProcessor
import logging



from monitor.models import (
    ConfiguracaoColeta, 
    LogExecucao, 
    Documento, 
    NormaVigente,
    RelatorioGerado
)


logger = logging.getLogger(__name__)

@shared_task(bind=True)
def executar_coleta_completa(self):
    """
    Tarefa principal que executa todo o fluxo de coleta e processamento
    """
    log_execucao = None
    config = ConfiguracaoColeta.objects.first()
    
    try:
        # Criar log de execução
        log_execucao = LogExecucao.objects.create(
            tipo_execucao='COMPLETO',
            status='PROCESSANDO',
            usuario=None,  # Pode ser substituído por usuário se autenticado
            mensagem="Iniciando fluxo completo de coleta"
        )
        
        # 1. Coleta do Diário Oficial
        logger.info("Iniciando coleta do Diário Oficial")
        diario_scraper = DiarioOficialScraper(max_docs=config.max_documentos if config else 10)
        documentos_coletados = diario_scraper.iniciar_coleta()
        
        # 2. Processamento dos documentos
        logger.info("Iniciando processamento dos documentos")
        processor = PDFProcessor()
        resultado_processamento = processor.processar_todos_documentos()
        
        # 3. Verificação na SEFAZ
        logger.info("Iniciando verificação na SEFAZ")
        integrador = IntegradorSEFAZ()
        normas_verificadas = integrador.verificar_documentos_nao_verificados()
        
        # Atualizar log com sucesso
        log_execucao.status = 'SUCESSO'
        log_execucao.data_fim = timezone.now()
        log_execucao.documentos_coletados = len(documentos_coletados)
        log_execucao.documentos_processados = resultado_processamento['sucesso']
        log_execucao.normas_verificadas = len(normas_verificadas)
        log_execucao.mensagem = "Fluxo completo executado com sucesso"
        log_execucao.save()
        
        # Enviar notificação por e-mail se configurado
        if config and config.email_notificacao:
            enviar_notificacao.delay(
                config.email_notificacao,
                log_execucao.id,
                'sucesso'
            )
        
        return {
            'status': 'success',
            'documentos_coletados': len(documentos_coletados),
            'documentos_processados': resultado_processamento['sucesso'],
            'normas_verificadas': len(normas_verificadas)
        }
        
    except Exception as e:
        logger.error(f"Erro na execução da coleta completa: {str(e)}", exc_info=True)
        
        if log_execucao:
            log_execucao.status = 'ERRO'
            log_execucao.data_fim = timezone.now()
            log_execucao.erro = str(e)
            log_execucao.traceback = self.request.get('traceback', '')
            log_execucao.save()
            
            if config and config.email_notificacao:
                enviar_notificacao.delay(
                    config.email_notificacao,
                    log_execucao.id,
                    'erro'
                )
        
        return {
            'status': 'error',
            'message': str(e)
        }

@shared_task
def verificar_coletas_programadas():
    """
    Verifica e executa coletas agendadas conforme configuração
    """
    try:
        config = ConfiguracaoColeta.objects.first()
        
        if not config or not config.ativa:
            logger.info("Coleta automática desativada ou não configurada")
            return False
        
        agora = timezone.now()
        
        if not config.proxima_execucao or agora >= config.proxima_execucao:
            logger.info("Executando coleta programada")
            
            # Atualizar horários de execução
            config.ultima_execucao = agora
            config.proxima_execucao = agora + timedelta(hours=config.intervalo_horas)
            config.save()
            
            # Executar coleta
            resultado = executar_coleta_completa.delay()
            
            return {
                'task_id': resultado.id,
                'proxima_execucao': config.proxima_execucao
            }
        
        logger.info(f"Próxima coleta programada para: {config.proxima_execucao}")
        return False
        
    except Exception as e:
        logger.error(f"Erro ao verificar coletas programadas: {str(e)}", exc_info=True)
        return False
    

@shared_task(name='monitor.tasks.gerar_relatorio_excel')
def gerar_relatorio_excel():
    logger.info("Gerando relatório Excel...")
    # Sua lógica de geração de relatório aqui
    return "Relatório gerado com sucesso"
    

@shared_task
def gerar_relatorio_contabil(data_inicio=None, data_fim=None, usuario_id=None):
    """
    Gera relatório contábil completo com os dados coletados
    """
    try:
        # Preparar nome do arquivo
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_contabil_{timestamp}.xlsx"
        caminho_arquivo = os.path.join(settings.MEDIA_ROOT, 'relatorios', nome_arquivo)
        
        # Criar diretório se não existir
        os.makedirs(os.path.dirname(caminho_arquivo), exist_ok=True)
        
        # Criar workbook
        wb = Workbook()
        ws_docs = wb.active
        ws_docs.title = "Documentos"
        
        # Adicionar cabeçalhos
        cabecalhos = [
            "ID", "Título", "Data Publicação", "Assunto", 
            "Processado", "Relevante", "Normas Relacionadas"
        ]
        ws_docs.append(cabecalhos)
        
        # Formatar cabeçalhos
        for cell in ws_docs[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Filtrar documentos
        documentos = Documento.objects.all()
        if data_inicio:
            documentos = documentos.filter(data_publicacao__gte=data_inicio)
        if data_fim:
            documentos = documentos.filter(data_publicacao__lte=data_fim)
        
        # Adicionar dados
        for doc in documentos.order_by('-data_publicacao'):
            ws_docs.append([
                doc.id,
                doc.titulo,
                doc.data_publicacao.strftime("%d/%m/%Y") if doc.data_publicacao else "",
                doc.assunto or "",
                "Sim" if doc.processado else "Não",
                "Sim" if doc.relevante_contabil else "Não",
                ", ".join([str(n) for n in doc.normas_relacionadas.all()])
            ])
        
        # Adicionar aba de normas
        ws_normas = wb.create_sheet(title="Normas")
        ws_normas.append([
            "Tipo", "Número", "Situação", "Última Verificação", 
            "Documentos Relacionados", "Fonte"
        ])
        
        # Formatar cabeçalhos
        for cell in ws_normas[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Adicionar normas
        normas = NormaVigente.objects.all().annotate(
            num_docs=Count('documentos')
        ).order_by('tipo', 'numero')
        
        for norma in normas:
            ws_normas.append([
                norma.get_tipo_display(),
                norma.numero,
                norma.get_situacao_display(),
                norma.data_verificacao.strftime("%d/%m/%Y %H:%M") if norma.data_verificacao else "",
                norma.num_docs,
                norma.fonte
            ])
        
        # Salvar arquivo
        wb.save(caminho_arquivo)
        
        # Registrar no banco de dados
        relatorio = RelatorioGerado.objects.create(
            tipo='CONTABIL',
            caminho_arquivo=caminho_arquivo.replace(settings.MEDIA_ROOT, ''),
            formato='XLSX',
            parametros={
                'data_inicio': data_inicio.isoformat() if data_inicio else None,
                'data_fim': data_fim.isoformat() if data_fim else None,
            },
            gerado_por_id=usuario_id
        )
        
        logger.info(f"Relatório contábil gerado: {caminho_arquivo}")
        
        return {
            'status': 'success',
            'relatorio_id': relatorio.id,
            'caminho': relatorio.caminho_arquivo.url if hasattr(relatorio.caminho_arquivo, 'url') else relatorio.caminho_arquivo
        }
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório contábil: {str(e)}", exc_info=True)
        return {
            'status': 'error',
            'message': str(e)
        }

@shared_task
def gerar_relatorio_mudancas(dias_retroativos=30, usuario_id=None):
    """
    Gera relatório de mudanças nas normas
    """
    try:
        integrador = IntegradorSEFAZ()
        mudancas = integrador.comparar_mudancas(dias_retroativos)
        
        # Preparar nome do arquivo
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_mudancas_{timestamp}.xlsx"
        caminho_arquivo = os.path.join(settings.MEDIA_ROOT, 'relatorios', nome_arquivo)
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mudanças nas Normas"
        
        # Adicionar cabeçalhos
        ws.append(["Tipo de Mudança", "Norma", "Detalhes"])
        
        # Formatar cabeçalhos
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Adicionar novas normas
        if mudancas['novas_normas']:
            ws.append(["Novas Normas Identificadas", "", ""])
            for norma in mudancas['novas_normas']:
                ws.append(["", norma, "Nova norma identificada no Diário Oficial"])
        
        # Adicionar normas revogadas
        if mudancas['normas_revogadas']:
            ws.append(["", "", ""])
            ws.append(["Normas Potencialmente Revogadas", "", ""])
            for item in mudancas['normas_revogadas']:
                ws.append(["", item['norma'], f"Última menção: {item['ultima_menção']}"])
        
        # Salvar arquivo
        wb.save(caminho_arquivo)
        
        # Registrar no banco de dados
        relatorio = RelatorioGerado.objects.create(
            tipo='MUDANCAS',
            caminho_arquivo=caminho_arquivo.replace(settings.MEDIA_ROOT, ''),
            formato='XLSX',
            parametros={
                'dias_retroativos': dias_retroativos,
            },
            gerado_por_id=usuario_id
        )
        
        logger.info(f"Relatório de mudanças gerado: {caminho_arquivo}")
        
        return {
            'status': 'success',
            'relatorio_id': relatorio.id,
            'caminho': relatorio.caminho_arquivo.url if hasattr(relatorio.caminho_arquivo, 'url') else relatorio.caminho_arquivo
        }
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório de mudanças: {str(e)}", exc_info=True)
        return {
            'status': 'error',
            'message': str(e)
        }

@shared_task
def verificar_normas_sefaz():
    """
    Verifica as normas junto à SEFAZ e atualiza seus status
    """
    try:
        integrador = IntegradorSEFAZ()
        
        # Verificar normas que nunca foram verificadas ou estão desatualizadas
        normas = NormaVigente.objects.filter(
            models.Q(data_verificacao__isnull=True) |
            models.Q(data_verificacao__lt=timezone.now()-timedelta(days=30))
        ).order_by('tipo', 'numero')
        
        total_normas = normas.count()
        normas_atualizadas = 0
        erros = 0
        
        for norma in normas:
            try:
                vigente, detalhes = integrador.verificar_vigencia_com_detalhes(norma.tipo, norma.numero)
                
                norma.situacao = 'VIGENTE' if vigente else 'REVOGADA'
                norma.data_verificacao = timezone.now()
                norma.detalhes = detalhes
                norma.save()
                
                normas_atualizadas += 1
            except Exception as e:
                logger.error(f"Erro ao verificar norma {norma}: {str(e)}")
                erros += 1
                continue
        
        logger.info(f"Verificação de normas concluída: {normas_atualizadas} atualizadas, {erros} erros")
        
        return {
            'status': 'success',
            'total_normas': total_normas,
            'normas_atualizadas': normas_atualizadas,
            'erros': erros
        }
        
    except Exception as e:
        logger.error(f"Erro na verificação de normas: {str(e)}", exc_info=True)
        return {
            'status': 'error',
            'message': str(e)
        }

@shared_task
def enviar_notificacao(email, log_execucao_id, tipo='sucesso'):
    """
    Envia e-mail de notificação sobre execuções do sistema
    """
    try:
        log_execucao = LogExecucao.objects.get(id=log_execucao_id)
        
        if tipo == 'sucesso':
            assunto = f"[Monitor] Coleta concluída - {log_execucao.data_inicio.strftime('%d/%m/%Y')}"
            template = 'emails/notificacao_sucesso.html'
        else:
            assunto = f"[Monitor] ERRO na coleta - {log_execucao.data_inicio.strftime('%d/%m/%Y')}"
            template = 'emails/notificacao_erro.html'
        
        contexto = {
            'log': log_execucao,
            'site_url': settings.SITE_URL,
        }
        
        mensagem_html = render_to_string(template, contexto)
        mensagem_texto = render_to_string(template.replace('.html', '.txt'), contexto)
        
        send_mail(
            subject=assunto,
            message=mensagem_texto,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            html_message=mensagem_html,
            fail_silently=False,
        )
        
        logger.info(f"Notificação enviada para {email}")
        return True
        
    except Exception as e:
        logger.error(f"Erro ao enviar notificação: {str(e)}", exc_info=True)
        return False

@shared_task
def processar_documento_individual(documento_id):
    """
    Processa um documento individual de forma assíncrona
    """
    try:
        documento = Documento.objects.get(id=documento_id)
        processor = PDFProcessor()
        resultado = processor.processar_documento(documento)
        
        return {
            'status': 'success',
            'documento_id': documento.id,
            'relevante': resultado,
            'normas_relacionadas': documento.normas_relacionadas.count()
        }
        
    except Exception as e:
        logger.error(f"Erro ao processar documento {documento_id}: {str(e)}", exc_info=True)
        return {
            'status': 'error',
            'message': str(e)
        }